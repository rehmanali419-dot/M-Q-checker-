import streamlit as st
import google.generativeai as genai
from PIL import Image
import pypdf
import docx
import openpyxl
import io

st.set_page_config(page_title="M&Q Audit System", layout="wide")

st.title("🛡️ M&Q PDF, Image & Document Audit System")
st.caption("Manufacturing & Quality Plan Inspector (PDF, JPG, PNG, DOCX, XLSX Support)")

api_key = st.text_input("🔑 Enter Gemini API Key", type="password")

if api_key:
    genai.configure(api_key=api_key)
    
    st.subheader("📄 Upload Audit Documents")
    
    # Upload all formats in one box
    uploaded_files = st.file_uploader(
        "Upload M&Q Documents / Images (PDF, JPG, PNG, DOCX, XLSX)", 
        type=["pdf", "jpg", "jpeg", "png", "docx", "xlsx"],
        accept_multiple_files=True
    )
    
    master_drawing = st.file_uploader(
        "Optional: Master Engineering Drawing (Image/PDF)", 
        type=["pdf", "jpg", "jpeg", "png"]
    )
    
    if st.button("🔍 Run Full M&Q Audit", type="primary"):
        if not uploaded_files:
            st.warning("Please upload at least one document or image to start the audit.")
        else:
            with st.spinner("Analyzing documents for dimensional errors and typing mistakes..."):
                try:
                    model = genai.GenerativeModel('gemini-1.5-flash')
                    contents = []
                    
                    for file in uploaded_files:
                        file_type = file.name.split('.')[-1].lower()
                        
                        if file_type in ['jpg', 'jpeg', 'png']:
                            image = Image.open(file)
                            contents.append(image)
                        elif file_type == 'pdf':
                            pdf_reader = pypdf.PdfReader(file)
                            text = ""
                            for page in pdf_reader.pages:
                                text += page.extract_text() or ""
                            contents.append(f"PDF Content ({file.name}):\n{text}")
                        elif file_type == 'docx':
                            doc = docx.Document(file)
                            text = "\n".join([p.text for p in doc.paragraphs])
                            contents.append(f"Word Document ({file.name}):\n{text}")
                        elif file_type == 'xlsx':
                            wb = openpyxl.load_workbook(file)
                            excel_text = ""
                            for sheet in wb.sheetnames:
                                ws = wb[sheet]
                                excel_text += f"\nSheet: {sheet}\n"
                                for row in ws.iter_rows(values_only=True):
                                    excel_text += " | ".join([str(cell) for cell in row if cell is not None]) + "\n"
                            contents.append(f"Excel Document ({file.name}):\n{excel_text}")

                    # Process optional drawing if provided
                    if master_drawing:
                        m_type = master_drawing.name.split('.')[-1].lower()
                        if m_type in ['jpg', 'jpeg', 'png']:
                            contents.append(Image.open(master_drawing))
                        elif m_type == 'pdf':
                            pdf_reader = pypdf.PdfReader(master_drawing)
                            m_text = "".join([p.extract_text() or "" for p in pdf_reader.pages])
                            contents.append(f"Master Drawing PDF:\n{m_text}")

                    prompt = """
                    You are an expert M&Q (Manufacturing & Quality) inspector. 
                    Carefully inspect the provided documents/images.
                    Cross-check all dimensional annotations, tolerances, sizes, and typing errors 
                    between sketch pages, specification tables, and requirement sheets.
                    
                    Provide a detailed discrepancy report highlighting:
                    1. Dimensional Errors / Size Mismatches (e.g., L4=722 vs L4=772)
                    2. Typing Errors or Missing Parameters
                    3. Recommendations for Correction
                    """
                    contents.append(prompt)
                    
                    response = model.generate_content(contents)
                    st.success("Audit Completed!")
                    st.markdown(response.text)
                    
                except Exception as e:
                    st.error(f"Error during audit: {str(e)}")
else:
    st.info("Please enter your Gemini API Key above to unlock.")
